import openpyxl
import time
wb = openpyxl.load_workbook('Final Consolidated RTE Requirements.xlsx')
sheets =  wb.get_sheet_names()
for name in sheets:
	sheet = wb.get_sheet_by_name(name)
	print(sheet , sheet.max_row)
	for i in range(2,sheet.max_row+1):
		tar = 'C'+str(i)
		src = 'D'+str(i)
		s = sheet[src].value
		if s!=None and s.startswith('G'):
			sheet[tar] = 'V'+sheet[tar].value[8:]
		if s!= None and s.startswith('C'):
			sheet[tar] = 'S'+sheet[tar].value[8:]
		print(sheet[tar].value+'\n')
wb.save('Res.xlsx')
                        
		
